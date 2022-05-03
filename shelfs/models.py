from django.db import models
from outils.common_imports import *

from outils.common import OdooAPI
from products.models import CagetteProducts
from inventory.models import CagetteInventory

import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from statistics import *


# Prefix for temp shelf inventory files
tmp_inv_file_prefix = 'temp/inventory_shelf_'

def as_text(value):
    """ Utils """
    return str(value) if value is not None else ""

class Shelf(models.Model):

    def __init__(self, id=0):
        """Init with odoo id."""
        self.id = int(id)
        self.o_api = OdooAPI()

    def get(self):
        res = {}
        try:
            c = [['id', '=', self.id]]
            f = []
            data = self.o_api.search_read('product.shelfs', c, f)
            res = data[0]
        except Exception as e:
            res['error'] = "Le rayon n'a pas pu être trouvé (" + str(e) + ")"
        return res

    def get_products(self):
        res = {}
        try:
            c = [['shelf_id', '=', self.id]]
            f = [
                'barcode',
                'name',
                'uom_id',
                'qty_available',
                'standard_price',
                'last_inv_delta',
                'last_inv_losses',
                'active'
            ]

            pdts = self.o_api.search_read('product.product', c, f)
            for p in pdts:
                for k, v in p.items():
                    if v is False:
                        p[k] = ''
            res['data'] = pdts
        except Exception as e:
            res['error'] = "Le rayon n'a pas pu être créé (" + str(e) + ")"
        return res

    def create(self, params):
        res = {}
        try:
            name = params.get('name')
            description = params.get('description')
            sort_order = params.get('sort_order')
            f = {'name': name, 'description': description, 'sort_order': sort_order}
            self.id = self.o_api.create('product.shelfs',f)
            f['id'] = self.id
            res = f
        except Exception as e:
            res['error'] = "Le rayon n'a pas pu être créé (" + str(e) + ")"
        return res

    def update(self, params):
        res = {}
        try:
            id = int(params.get('id'))
            name = params.get('name')
            description = params.get('description')
            sort_order = params.get('sort_order')
            f = {'name': name, 'description': description, 'sort_order': sort_order}

            result = self.o_api.update('product.shelfs', id, f)

            if result:
                c = [['id','=', id]]
                res = self.o_api.search_read('product.shelfs', c)
                res = res[0]
            else:
                res['error'] = "Le rayon n'a pas pu être mis à jour"
        except Exception as e:
            res['error'] = "Le rayon n'a pas pu être mis à jour (" + str(e) + ")"
        return res

    def update_shelf_with_inventory_data(self, params):
        res = {}
        f = {}
        if params['inventory_status'] == '' :
            f['inventory_status'] = 'step1_done'
        elif params['inventory_status'] == 'step1_done' :
            # Inventory all done
            f['inventory_status'] = ''
            f['date_last_inventory'] = date.today().strftime("%Y-%m-%d")

            if 'last_inventory_id' in params:
                f['last_inventory_id'] = params['last_inventory_id']
            if 'shelf_delta' in params:
                f['last_inv_delta_percentage'] = params['shelf_delta']
            if 'shelf_losses' in params:
                f['last_inv_losses_percentage'] = params['shelf_losses']

        res['update'] = self.o_api.update('product.shelfs', self.id, f)

        return res

    def update_products_with_inventory_data(self, products):
        done = []
        missed = []

        for p in products:
            try:
                f = {
                'last_inv_delta': p['delta'],
                'last_inv_losses': p['losses']
                }

                result = self.o_api.update('product.product', p['id'], f)
                done.append({'id': p['id'], 'update': result})
            except Exception as e:
                missed.append({'id': p['id'], 'update': False})

        return {'missed': missed, 'done': done}

    def delete(self, params):
        res = {}
        try:
            shelf_id = params.get('id')
            res = self.o_api.execute('product.shelfs', 'unlink', [shelf_id])
        except Exception as e:
            coop_logger.error("Rayon, delete : %s, %s", str(e), str(params))
            res['error'] = "Le rayon n'a pas pu être détruit"
        return res


    def _get_pdts_from_barcodes(self, barcodes):
        c = [['barcode', 'in', barcodes]]
        f = ['barcode']
        return self.o_api.search_read('product.product', c, f)

    def add_products_by_barcodes(self, barcodes):
        res = {}
        try:
            #  TODO : send directly barcodes as params (to long to test /debug for now)
            pids = []
            barcodes = list(map(str, barcodes))
            # Get bc as should be stored in Odoo
            bc_map = CagetteProducts.get_fixed_barcode_correspondance(barcodes)

            p_res = self._get_pdts_from_barcodes(list(bc_map.values()))
            if (p_res):
                found_bc = []
                for p in p_res:
                    pids.append(p['id'])
                    found_bc.append(p['barcode'])

                if (len(p_res) != len(barcodes)):
                    res['missing'] = list(set(barcodes) ^ set(found_bc))
                params = {'shelf_id': self.id, 'pids': pids}
                res['added'] = self.o_api.execute('product.shelfs', 'add_products', params)
            else:
                # None of the list have been found
                res['msg'] = "No product found"
        except Exception as e:
            coop_logger.error("Rayons, add_products_by_barcodes : %s, %s", str(e), str(barcodes))
            res['error'] = "L'enregistrement n'a pas pu se réaliser"
        return res

    def update_last_product_added_date(self):
        res = {}
        today = date.today().strftime("%Y-%m-%d")
        f = {'date_last_product_added': today}

        try:
            res["update"] = self.o_api.update('product.shelfs', self.id, f)
        except Exception as e:
            res['error'] = str(e)

        return res


    def save_tmp_inventory(self, inventory_data):
        """Save inventory data in a json temp file"""
        res = {}
        try:
            filename = tmp_inv_file_prefix + str(self.id) + '.json'
            with open(filename, 'w') as outfile:
                json.dump(inventory_data, outfile)

            res['file_saved'] = True
        except Exception as e:
            res['error'] = str(e)

        return res

    def remove_tmp_inventory(self):
        """Remove json temp inventory file"""
        try:
            filename = tmp_inv_file_prefix + str(self.id) + '.json'
            os.remove(filename)

            lockfilename = tmp_inv_file_prefix + str(self.id) + '.lock'
            try:
                os.remove(lockfilename)
            except Exception as e:
                pass

            return True
        except Exception as e:
            return False

    def get_full_inventory_data(self, inventory_data) :
        """
            Get data from json file of inventory's first step,
            and concatenate product qties of both steps.

            Also check for duplicates.
                If found, concatenate qties and remove duplicate.
        """
        # get saved data from first step
        first_inventory = None
        try:
            filename = tmp_inv_file_prefix + str(self.id) + '.json'
            with open(filename) as json_file:
                first_inventory = json.load(json_file)
        except Exception as e:
            coop_logger.error("Unable to process first step file : %s", e)
            import errno
            raise FileExistsError(errno.ENOENT, os.strerror(errno.ENOENT), filename)

        lockfilename = tmp_inv_file_prefix + str(self.id) + '.lock'

        # Look for lock file: if exists, first step file is being processed so stop here
        try:
            with open(lockfilename) as lock_file:
                return {'error': 'First step file busy', 'busy': True}
        except Exception as e:
            pass

        # Verification passed, create the lock file to indicate first step file is being processed
        try:
            with open(lockfilename, 'w') as lock_file:
                json.dump({}, lock_file)
        except Exception as e:
            coop_logger.error("Unable to create lock file : %s", e)

        if first_inventory:
            # if poducts in saved data
            if 'products' in first_inventory :
                # for each product in incoming inventory data
                for p in inventory_data['products'] :
                    # find product in saved data
                    matching_product = None
                    mp_index = -1
                    for i, p2 in enumerate(first_inventory['products'], start=0):
                        if p['id'] == p2['id']:
                            mp_index = i
                            matching_product = p2

                    # concatenate qtys if found
                    if mp_index != -1:
                        qty = float(p['qty']) + float(matching_product['qty'])
                        p['qty'] = qty

                        # pop product in first_inventory products list
                        first_inventory['products'].pop(mp_index)

            # if products remain in first step data, add it to inventory data
            if len(first_inventory['products']) > 0 :
                inventory_data['products'] = inventory_data['products'] + first_inventory['products']

            # Add user comments from step 1
            if 'user_comments' in first_inventory:
                inventory_data['user_comments_step1'] = first_inventory['user_comments']
        else:
            inventory_data['error'] = "Missing first step data"

        # Make another pass through data to remove duplicates
        safe_from_duplicates = []
        for p in inventory_data['products']:
            found = False
            for p2 in safe_from_duplicates:
                if p['id'] == p2['id']:
                    # Concatenate qties for duplicates
                    qty = float(p['qty']) + float(p2['qty'])
                    p2['qty'] = qty
                    found = True

            if not found:
                safe_from_duplicates.append(p)

        inventory_data['products'] = safe_from_duplicates

        return inventory_data

    def save_inventory_report(self, inventory_data) :
        """Save a report of differences found after a shelf inventory"""
        res = {}
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Rapport d'inventaire"

            ref_count = 0
            units_count_kg = 0
            units_count_unit = 0
            initial_stock_unit = 0
            initial_stock_kg = 0

            initial_shelf_value = 0
            inventory_value = 0

            shelf_delta_abs_unit = 0
            shelf_delta_rel_unit = 0

            shelf_delta_abs_kg = 0
            shelf_delta_rel_kg = 0

            ws.append( ['Rapport d\'inventaire'] )
            ws.append( ['Rayon (Numéro) :', inventory_data['shelf_name'] + ' (' + str(inventory_data['shelf_num']) + ')'] )
            ws.append( ['Date :', inventory_data['date'].strftime("%d/%m/%Y")] )
            ws.append( [] )
            ws.append( ['Nom de l\'article', 'UoM', 'Stock théorique initial', 'Stock final', 'Delta (Qté)', 'Pertes (€)'] )
            titles_row = ws.max_row

            for p in inventory_data['products'] :
                p['delta'] = delta = float(p['qty']) - float(p['qty_available'])
                p['losses'] = losses = delta * float(p['standard_price'])
                ws.append( [p['name'], p['uom_id'][1], round(p['qty_available'], 2), round(p['qty'], 2), round(delta, 2), round(losses, 2)] )

                # Count references in shelf after inventory
                if p['qty'] > 0:
                    ref_count += 1

                # Calculate value in shelf after inventory
                inventory_value += float(p['qty']) * float(p['standard_price'])
                # And before as well
                initial_shelf_value += float(p['qty_available']) * float(p['standard_price'])

                # For products sold by unit
                if p['uom_id'][0] == 1:
                    # Count initial stock
                    initial_stock_unit += float(p['qty_available'])

                    # Count units in shelf
                    units_count_unit += float(p['qty'])

                    # Global delta in absolute value (found + missing units count)
                    shelf_delta_abs_unit += abs(delta)

                    # Global delta in relative value (found - missing units count)
                    shelf_delta_rel_unit += delta

                # For products sold by kg
                else:
                    initial_stock_kg += float(p['qty_available'])
                    units_count_kg += float(p['qty'])
                    shelf_delta_abs_kg += abs(delta)
                    shelf_delta_rel_kg += delta

            # END FOR

            # Calculate shelf delta in %
            if initial_stock_unit > 0 :
                shelf_delta_rel_unit_percent = (shelf_delta_rel_unit / initial_stock_unit) * 100
            else :
                shelf_delta_rel_unit_percent = 0

            if initial_stock_kg > 0 :
                shelf_delta_rel_kg_percent = (shelf_delta_rel_kg / initial_stock_kg) * 100
            else :
                shelf_delta_rel_kg_percent = 0

            # Losses in shelf (initial shelf value - final value)
            shelf_losses = initial_shelf_value - inventory_value

            # Shelf losses in % : shelf_losses / initial stock
            if initial_shelf_value > 0 :
                shelf_losses_percent = (shelf_losses / initial_shelf_value) * 100
            else :
                shelf_losses_percent = 0

            ws.append( [] )

            ws.append( ['Nombre total de références :', ref_count] )
            ws.append( ['Quantité totale de produits au KG :', round(units_count_kg, 2)] )
            ws.append( ['Quantité totale de produits à l\'unité :', round(units_count_unit, 2)] )
            ws.append( [] )
            ws.append( ['Valeur d\'inventaire du rayon :', round(inventory_value, 2)] )

            ws.append( [] )
            ws.append( ['Deltas du rayon'] )
            ws.append( ['Pour les produits à l\'UNITÉ'] )
            ws.append( ['   Delat absolu : ', round(shelf_delta_abs_unit, 2), '(= qté d\'unités trouvées + qté d\'unités disparues)'] )
            ws.append( ['   Delta relatif : ', round(shelf_delta_rel_unit, 2), '(= qté d\'unités trouvées - qté d\'unités disparues)'] )
            ws.append( ['   Delta relatif en % : ', round(shelf_delta_rel_unit_percent, 2), '(= delta relatif / stock total théorique initial)'] )
            ws.append( ['Pour les produits au KG'] )
            ws.append( ['   Delta absolu : ', round(shelf_delta_abs_kg, 2)] )
            ws.append( ['   Delta relatif : ', round(shelf_delta_rel_kg, 2)] )
            ws.append( ['   Delta relatif en % : ', round(shelf_delta_rel_kg_percent, 2)] )

            ws.append( [] )
            ws.append( ['Pertes pour le rayon : ', round(shelf_losses, 2), '(= valeur totale avant comptage - valeur totale d\'inventaire)'] )
            ws.append( ['Pertes en % du rayon : ', round(shelf_losses_percent, 2), '(= pertes du rayon / valeur totale avant comptage)'] )

            # User comments
            ws.append( [] )
            if 'user_comments_step1' not in inventory_data or inventory_data['user_comments_step1'] == '':
                inventory_data['user_comments_step1'] = 'Aucun'

            ws.append( ['Problèmes survenus durant le comptage en rayon :', inventory_data['user_comments_step1']] )
            # Merge cells for comments
            merge_begin = ws.max_row
            merge_end = ws.max_row+3
            ws.append( [] )
            ws.append( [] )
            ws.append( [] )
            ws.merge_cells(start_row=merge_begin, start_column=1, end_row=merge_end, end_column=1)
            ws.merge_cells(start_row=merge_begin, start_column=2, end_row=merge_end, end_column=6)
            # Styling merged cells
            top_left_cell = ws['A'+str(merge_begin)]
            top_left_cell.alignment = Alignment(vertical="top")
            top_left_cell.font = Font(bold=True)
            top_left_cell = ws['B'+str(merge_begin)]
            top_left_cell.alignment = Alignment(vertical="top")

            if 'user_comments' not in inventory_data or inventory_data['user_comments'] == '':
                inventory_data['user_comments'] = 'Aucun'
            ws.append( ['Problèmes survenus durant le comptage en réserve :', inventory_data['user_comments']] )
            merge_begin = ws.max_row
            merge_end = ws.max_row+3
            ws.append( [] )
            ws.append( [] )
            ws.append( [] )
            ws.merge_cells(start_row=merge_begin, start_column=1, end_row=merge_end, end_column=1)
            ws.merge_cells(start_row=merge_begin, start_column=2, end_row=merge_end, end_column=6)
            top_left_cell = ws['A'+str(merge_begin)]
            top_left_cell.alignment = Alignment(vertical="top")
            top_left_cell.font = Font(bold=True)
            top_left_cell = ws['B'+str(merge_begin)]
            top_left_cell.alignment = Alignment(vertical="top")


            # Styling

            # Columns size
            for column_cells in ws.columns:
                # "Auto fit" the first column width to content
                length = max(len(as_text(cell.value)) for cell in column_cells)

                # For other columns than the first, set size
                if column_cells[3].column_letter != "A" :
                    length = 20

                ws.column_dimensions[column_cells[3].column_letter].width = length

            # First column Bold
            col = ws.column_dimensions['A']
            col.font = Font(bold=True)

            # Make that line bold
            row = ws.row_dimensions[titles_row]
            row.font = Font(bold=True)


            # Save file
            report_file_name = "inventaire_rayon"+str(round(inventory_data['shelf_num'], 2))+"_"+inventory_data['date'].strftime("%Y-%m-%d")+".xlsx"

            wb.save(filename=report_file_name)

            # Attach report to Odoo inventory instance
            inventory = CagetteInventory(inventory_data['inventory_id'])
            inventory.attach_file(report_file_name)

            res['msg'] = 'Rapport enregistré.'

            # Store some inventory data in shelf & products tables
            res['products'] = inventory_data['products']
            res['shelf_losses'] = round(shelf_losses_percent, 2)
            # For delta, keep only one
            if shelf_delta_rel_kg_percent == 0:
                res['shelf_delta'] = round(shelf_delta_rel_unit_percent, 2)
            else:
                res['shelf_delta'] = round(shelf_delta_rel_kg_percent, 2)
        except Exception as e:
            res =  'Erreur lors de l\'enregistrement du rapport : ' + str(e)

        return res

class Shelfs(models.Model):

    def get_all():
        res = []
        try:
            api = OdooAPI()
            res = api.execute('product.shelfs', 'get', {})
        except Exception as e:
            coop_logger.error("Rayons, get_all : %s", str(e))
        return res

    @staticmethod
    def get_shelfs_sortorder(shelf_ids=[]):
        """Get all shelves sort_order"""
        res = []
        try:
            api = OdooAPI()
            if len(shelf_ids) == 0:
                c = []
            else:
                c = [['id', 'in', shelf_ids]]
            f = ['id', 'sort_order']
            res = api.search_read('product.shelfs', c, f)
        except Exception as e:
            coop_logger.error("Rayons, get_shelfs_sortorder : %s", str(e))
        return res
