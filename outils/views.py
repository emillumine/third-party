# coding: utf-8
"""Root main page."""

from outils.common_imports import *
from outils.for_view_imports import *

from django.views.generic import View

from django.views.decorators.csrf import csrf_exempt

from .common import OdooAPI

from .forms import OdooEntityFieldsForm
from .forms import ExportComptaForm
from outils.lib.compta import *
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from orders.models import Orders


def test_compta(request):
    generate_quadratus_compatible_file('bidon')
    return HttpResponse('ok')


def index(request):
    u"""Page de présentation."""
    template = loader.get_template('root.html')
    context = {'form': '', 'title': 'Outils La Cagette'}
    response = HttpResponse(template.render(context, request))
    return response

@csrf_exempt
def log_js_error(request):
    """TODO : Prevent DOS by filling the log file."""
    try:
        paris_tz = pytz.timezone('Europe/Paris')
        file = open('outils/js_errors.log', 'a')

        line = [str(datetime.datetime.now(tz=paris_tz)),
                request.POST.get('module'),
                request.META.get('HTTP_USER_AGENT', ''),
                request.POST.get('error')]
        file.write("\t".join(line) + "\n")
        file.close()
    except Exception as e:
        coop_logger.error("log_js_error : %s", str(e))

    return HttpResponse('ok')


def data(request, fn):
    """Return file octet stream, located in data folder.
    Useful when frontend webserver has not direct access to django filesystem
    """
    from django.views.static import serve
    import os.path
    filepath = 'data/'+ fn
    return serve(request, os.path.basename(filepath), os.path.dirname(filepath))

def entity_example(request):
    # 536521 : dernier id accessible pour account.move.line !!
    entity = request.GET.get("entity", "")
    fields = request.GET.get("fields", [])
    id = request.GET.get("id", 0)
    api = OdooAPI()
    cond = []
    if id != '' and int(id) > 0:
        cond.append(['id', '=', int(id)])
    example = api.search_read(entity, cond, fields.split('|'), 20, 0, 'id DESC')
    # commenté à cause de shift.template : AttributeError: \'shift.template\' object has no attribute \'_compute_seats\'
    #example = api.search_read(entity, cond, [], 20)
    # example = api.search_read('shift.template.registration.line', [['registration_id','=',1850]], [], 20)
    response = JsonResponse({'example': example})
    return response

class FieldsView(View):
    u"""Permet de visualiser, pour chaque objet les champs associés."""

    def get(self, request, *args, **kwargs):
        u"""Affichage du formulaire pour récupérer le nom de l'objet"""
        template = loader.get_template('common/fields.html')
        context = {'form': OdooEntityFieldsForm(),
                   'title': 'Attributs accessibles'}

        return HttpResponse(template.render(context, request))

    def post(self, request, *args, **kwargs):
        u"""Nous allons retourner les attributs Odoo correspondants."""
        entity = request.POST.get("entity", "")
        fields = []
        if getattr(settings, 'APP_ENV', "prod") == "dev" and len(entity) > 1:
            api = OdooAPI()
            fields = api.get_entity_fields(entity)
        template = loader.get_template('common/entity_fields.html')
        context = {'fields': fields, 'entity': entity,
                   'title': 'Odoo Entity fields'}

        return HttpResponse(template.render(context, request))

class ExportCompta(View):
    u"""Traitement du fichier obtenu par export compta d'Odoo."""



    def get(self, request, *args, **kwargs):
        u"""Affichage du formulaire pour récupérer le fichier CSV Odoo."""
        if request.GET.get('phase', None) is None:
            template = loader.get_template('outils/export_compta.html')

            context = {'form': ExportComptaForm(),
                       'title': 'Export compta'}
            if hasattr(settings, 'EXPORT_COMPTA_FORMAT'):
                context['final_format'] = settings.EXPORT_COMPTA_FORMAT

            response = HttpResponse(template.render(context, request))
        else:
            val = None
            try:
                export_id = (int)(request.GET.get('export_id'))

                if request.GET.get('phase') == '1':
                    val = generate_account_export_report(export_id)
                if request.GET.get('phase') == '2':
                    final_format = request.GET.get('final_format', 'csv')
                    val = get_odoo_account_export_report(export_id, final_format)

                response = JsonResponse({'response': val})
            except Exception as e:
                val = str(e)
                response = JsonResponse({'error': val})
        return response



    def post(self, request, *args, **kwargs):
        u"""Month and year are given: a file is generated to fit with Arithmetic software or Quadra"""

        response = generate_odoo_export_file(request.POST.get('from'), request.POST.get('to'))

        return JsonResponse({'response': response})


class ExportPOS(View):
    u"""Export lié POS d'Odoo."""

    def get(self, request, *args, **kwargs):
        u"""Affichage du formulaire pour récupérer le fichier d'Export Odoo."""
        template = loader.get_template('outils/export_pos.html')
        context = {'title': 'Export POS'}

        return HttpResponse(template.render(context, request))

    def __ca_sessions_ng(self, mois):
        import re
        debut = time.time()
        api = OdooAPI()
        res = api.execute('lacagette.pos_payments_export', 'get_pos_payments', {'month' : mois})
        if not ('sessions' in res):
            return HttpResponse('Aucune session trouvée')
        tf_ym = "%Y-%m-%d %H:%M:%S"
        req_month = time.strptime(mois, "%Y-%m")
        kept_sessions_id = []
        totals = {}
        details_lines = []
        for s in res['sessions']:
            if 'min' in s['mm_dates']:
                """
                 s['mm_dates']['min'] and s['mm_dates']['max'] could be formatted with milliseconds
                 i.e 2020-12-12 12:38:58.136 (rescue Session)
                 Thus, .xxx has to be removed
                """
                s['mm_dates']['min'] = re.sub(r'\.[0-9]+', '', s['mm_dates']['min'])
                s['mm_dates']['max'] = re.sub(r'\.[0-9]+', '', s['mm_dates']['max'])
                min_date = time.strptime(s['mm_dates']['min'], tf_ym)
                max_date = time.strptime(s['mm_dates']['max'], tf_ym)

                if min_date.tm_mon == max_date.tm_mon and min_date.tm_mon == req_month.tm_mon:
                    y = str(min_date.tm_year)
                    m = min_date.tm_mon
                    d = min_date.tm_mday
                    if m < 10:
                        m = '0' + str(m)
                    else:
                        m = str(m)
                    if d < 10:
                        d = '0' + str(d)
                    else:
                        d = str(d)
                    kept_sessions_id.append(s['id'])
                    key = y + '-' + m + '-' + d
                    if not (key in totals):
                        totals[key] = {'CB': 0,
                                       'CSH': 0,
                                       'CHQ': 0,
                                       'CB_DEJ': 0,
                                       'CHQ_DEJ': 0,
                                       'MonA': 0,
                                       'TOTAL': 0}
                    sub_total = 0
                    cb = chq = csh = cbd = chqd = mona = 0
                    coop_logger.info("payments : %s", s['payments'])
                    for p in s['payments']:
                        # p['name'] is a sequence generated string
                        # Test order is important as CHEQDEJ contains CHEQ for ex.
                        # p['journal'] could be used but easier to change in Odoo interface
                        sub_amount = round(p['total_amount'], 2)
                        if 'CSH' in p['name']:
                            csh = sub_amount
                        elif 'CHEQDEJ' in  p['name']:
                            chqd = sub_amount
                        elif 'CHEQ' in p['name']:
                            chq = sub_amount
                        elif 'CBDEJ' in p['name']:
                            cbd = sub_amount
                        elif 'CB' in p['name']:
                            cb = sub_amount
                        elif 'MonA' in p['name'] or 'MonA' in p['journal']:
                            mona = sub_amount
                        sub_total += sub_amount
                    totals[key]['CB'] += cb
                    totals[key]['CSH'] += csh
                    totals[key]['CHQ'] += chq
                    totals[key]['CB_DEJ'] += cbd
                    totals[key]['CHQ_DEJ'] += chqd
                    totals[key]['MonA'] += mona
                    totals[key]['TOTAL'] += round(sub_total, 2)
                    details_lines.append([mois, s['mm_dates']['min'], s['mm_dates']['max'], s['caisse'], s['name'],
                                         cb, csh, chq, cbd, chqd, mona, sub_total])
        wb = Workbook()
        ws1 = wb.create_sheet("Totaux " + mois, 0)
        ws2 = wb.create_sheet("Détails " + mois, 1)
        ws1.append(['date', 'CB', 'CSH', 'CHQ', 'CB_DEJ', 'CHQ_DEJ', 'MonA', 'Total'])
        for day in sorted(totals):
            cb = totals[day]['CB']
            csh = totals[day]['CSH']
            chq = totals[day]['CHQ']
            cbd = totals[day]['CB_DEJ']
            chqd = totals[day]['CHQ_DEJ']
            mona = totals[day]['MonA']
            total = totals[day]['TOTAL']
            ws1.append([day, cb, csh, chq, cbd, chqd, mona, total])
        ws2.append(['mois', 'min_date', 'max_date', 'Caisse', 'session', 'CB', 'CSH','CHQ', 'CB_DEJ', 'CHQ_DEJ', 'MonA', 'total'])
        for row in details_lines:
            ws2.append(row)
        wb_name = 'export_sessions__' + mois + '.xlsx'
        # wb.save('/tmp/' + wb_name)
        response = HttpResponse(content=save_virtual_workbook(wb),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=' + wb_name
        return response

    def __ca_sessions(self, mois):
        api = OdooAPI()

        cond = [['stop_at', 'like', mois], ['state', '=', 'closed']]
        fields = ['stop_at', 'statement_ids', 'display_name', 'total_amount', 'order_ids', 'config_id']
        sessions = api.search_read('pos.session', cond, fields, 100000)
        statement_ids = []
        order_ids = []
        for s in sessions:
            statement_ids += s['statement_ids']
            order_ids += s['order_ids']
            s['min_date'] = '2970-01-01 00:00:00'
            s['max_date'] = '1970-01-01 00:00:00'
        # Verifions d'abord que toutes les dates de vente d'une session sont
        # toutes du même jour

        cond = [['order_id', 'in', order_ids]]
        fields = ['create_date', 'order_id']
        pol = api.search_read('pos.order.line', cond, fields, 100000)
        tf = "%Y-%m-%d %H:%M:%S"
        for l in pol:
            l_date = time.strptime(l['create_date'], tf)
            for s in sessions:
                s_min_date = time.strptime(s['min_date'], tf)
                s_max_date = time.strptime(s['max_date'], tf)
                for oi in s['order_ids']:
                    if (oi == l['order_id'][0]):
                        if (l_date < s_min_date):
                            s['min_date'] = l['create_date']
                        if (l_date > s_max_date):
                            s['max_date'] = l['create_date']

        cond = [['statement_id', 'in', statement_ids]]
        fields = ['pos_statement_id', 'statement_id', 'amount']
        payments = api.search_read('account.bank.statement.line', cond, fields, 100000)
        payments_statements = {}
        for p in payments:
            if not (p['statement_id'][0] in payments_statements):
                payments_statements[p['statement_id'][0]] = {'nom': p['statement_id'][1], 'total' : 0.00}
            payments_statements[p['statement_id'][0]]['total'] += round(p['amount'],2)
        res = []
        for s in sessions:
            line = s
            for p in s['statement_ids']:
                try:
                    line[payments_statements[p]['nom']] =\
                        round(payments_statements[p]['total'], 2)
                except:
                    # line['erreur_statement_id'] = p
                    # Aucun reglement avec le moyen de paiement
                    pass
            del s['statement_ids']
            line['cheval'] = False
            try:
                if (time.strptime(s['min_date'], "%Y-%m-%d") !=
                   time.strptime(s['max_date'], "%Y-%m-%d")):
                    line['cheval'] = True
            except:
                pass
            res.append(line)
        wb = Workbook()
        ws1 = wb.create_sheet("Totaux " + mois, 0)
        ws2 = wb.create_sheet("Détails " + mois, 1)
        ws2.append(['mois', 'min_date', 'max_date', 'Caisse', 'session', 'CB', 'CSH', 'CHQ', 'total'])

        totals = {}
        for r in res:
            if (r['total_amount'] > 0):
                # d = time.strptime(r['min_date'], tf)
                # date = str(d.tm_mday) + '-' + str(d.tm_mon) + '-' + str(d.tm_year)
                date, hours = r['min_date'].split(' ')
                total = round(r['total_amount'], 2)
                cb = csh = chq = 0
                caisse = r['config_id'][1]

                for key in r.keys():
                    if ('CB' in key):
                        cb = r[key]
                    if ('CSH' in key):
                        csh = r[key]
                    if ('CHEQ' in key):
                        chq = r[key]
                if not (date in totals):
                    totals[date] = {'CB': cb, 'CSH': csh, 'CHQ': chq, 'TOTAL': total}
                else:
                    totals[date]['CB'] += cb
                    totals[date]['CSH'] += csh
                    totals[date]['CHQ'] += chq
                    totals[date]['TOTAL'] += total
                line = [mois, r['min_date'], r['max_date'], caisse, r['display_name'],
                        cb, csh, chq, total]
                # writer.writerow(line)
                ws2.append(line)
        ws1.append(['date', 'CB', 'CSH', 'CHQ', 'Total'])

        for day in sorted(totals):
            cb = totals[day]['CB']
            csh = totals[day]['CSH']
            chq = totals[day]['CHQ']
            total = totals[day]['TOTAL']
            ws1.append([day, cb, csh, chq, total])
        wb_name = 'export_sessions__' + mois + '.xlsx'
        # wb.save('/tmp/' + wb_name)
        response = HttpResponse(content=save_virtual_workbook(wb),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=' + wb_name
        return response

    def post(self, request, *args, **kwargs):
        u"""Traitement formulaire."""
        month = request.POST.get("month", "")
        response = HttpResponse("Impossible de récupérer le fichier.")
        if len(month) == 0:
            import datetime
            month = datetime.date.today().strftime("%Y-%m")
        else:
            # verifions la validité du mois
            d_elts = month.split('-')
            if len(d_elts) == 2:
                try:
                    y = int(d_elts[0])
                    m = int(d_elts[1])
                    if (y < 100 * m):
                        month = None
                except:
                    month = None
            else:
                month = None

        if not (month is None):
            response = self.__ca_sessions_ng(month)
        return response

class ExportOrders(View):
    def get(self, request, *args, **kwargs):
        u"""Display form"""
        template = loader.get_template('outils/export_orders.html')
        context = {'title': 'Export Commandes Réceptionnées'}

        return HttpResponse(template.render(context, request))

    def post(self, request, *args, **kwargs):
        u"""Generate orders export between two dates"""
        date_from = request.POST.get('from')
        date_to = request.POST.get('to')

        orders = Orders.get_orders_between_dates(date_from, date_to)

        if "error" in orders:
            error = "Une erreur est survenue, merci de contacter le service informatique."
            return JsonResponse({'erreur': error, 'details': orders["error"]})

        try:
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Commandes réceptionnées"
            ws1.append(['Fournisseur', 'Réf commande', 'Statut', 'Montant HT', 'Montant Total', 'Date réception'])
            for order in orders["data"]:
                supplier_name = order['supplier_name']
                id_po = order['id_po']
                amount_untaxed = order['amount_untaxed']
                amount_total = order['amount_total']

                if order['state'] == "purchase":
                    state = "Commande fournisseur"
                elif order['state'] == "done":
                    state = "Terminé"
                else:
                    state = order['state']

                date_done_obj = datetime.datetime.strptime(order['date_done'], '%Y-%m-%d %H:%M:%S')
                date_done = date_done_obj.strftime("%d/%m/%Y")

                ws1.append([supplier_name, id_po, state, amount_untaxed, amount_total, date_done])

            wb_name = 'export_orders_' + date_from + '_' + date_to + '.xlsx'
            response = HttpResponse(content=save_virtual_workbook(wb),
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=' + wb_name

            return response
        except Exception as e:
            error = "Une erreur est survenue, merci de contacter le service informatique."
            coop_logger.error("Erreur export_orders : %s", str(e))
            return JsonResponse({'erreur': error, 'details': str(e)})
