# coding: utf-8
import sys, getopt, os
sys.path.append(os.path.abspath('../..'))
from outils.common import OdooAPI
from outils.config import COOP_BARCODE_RULE_ID
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import datetime
# from django.conf import settings

from openpyxl.utils.exceptions import InvalidFileException


def main():
    """ Pass file as script arg """
    if len(sys.argv) < 2:
        print("Il faut renseigner le chemin du fichier contenant les données des rattachés (format Excel).")
        exit(2)

    api = OdooAPI()
    data_file = sys.argv[1]

    try:
        wb = load_workbook(data_file)
    except FileNotFoundError:
        print("Fichier introuvable.")
        exit(2)
    except InvalidFileException:
        print("Le fichier fourni est invalide, il doit être au format Excel (.xlsx,.xlsm,.xltx,.xltm)")
        exit(2)

    while True:
        try:
            worksheet_number = input(f'Numéro de la feuille contenant les données (entre 0 et {len(wb.worksheets)-1}) : ')
            worksheet_number = int(worksheet_number)
            ws = wb.worksheets[worksheet_number]
        except ValueError:
            print("Veuillez rentrer un numéro entier.")
            continue
        except IndexError:
            print("Cette feuille n'existe pas.")
            continue
        else:
            break

    columns_valid = input("""
Les colonnes doivent être les suivantes :
A: active*
B: barode_base*
C: name*
D: (ignorée)
E: date_inclusion
F: email
G: birthdate
H: mobile
I: phone
J: street
K: street2
L: zip
M: city
N: sex
O: (ignorée)
P: parent_member_id*

(* ces champs doivent être renseignés)
Vous confirmez ? (O/n) """)

    while True:
        if (columns_valid == 'n' or columns_valid == 'N'):
            print("Veuillez formatter correctement le fichier avant de continuer !")
            exit()
        elif (columns_valid == 'o' or columns_valid == 'O' or columns_valid== ''):
            break
        else: 
            columns_valid = input("Vous confirmez ? (O/n) ")
            continue

    users = []
    has_error = False

    for row in ws.iter_rows(min_row=2, values_only=True):
        # active
        if row[0] is None or row[0] is False or row[0] == "=FALSE()" or row[0] == "=false":
            active = False 
        else:
            active = True

        # If line is not empty (mandatory field check)
        if row[2] is not None:
            user = {
                "is_member": False,
                "is_associated_people": True,
                "active": active,
                "barcode_rule_id": COOP_BARCODE_RULE_ID,
                "name": row[2],
                "parent_id": int(row[15]),  # for development, override with local existing member id
            }

            if row[1] is not None and row[1] != "NON":
                try:
                    user["barcode_base"] = int(row[1])
                except Exception:
                    print(f"[Mauvais format du champ 'barcode_base' pour l'utilisateur '{row[2]}' (Attendu : nombre entier)")
                    has_error = True

            if row[4] is not None:
                user["comment"] = f"Date d'inclusion : {row[4].date().strftime('%d/%m/%Y')}"

            if row[5] is not None:
                user["email"] = row[5]

            if row[6] is not None:
                user["birthdate"] = str(row[6].date())

            if row[7] is not None:
                try:
                    mobile = str(row[7]).replace(" ", "")
                    if mobile[0] != '+' and mobile[0] != '0':
                        mobile = '0' + mobile
                    user["mobile"] = mobile
                except Exception:
                    print(f"[Mauvais format du champ 'mobile' pour l'utilisateur '{row[2]}'")
                    has_error = True

            if row[8] is not None:
                try:
                    phone = str(row[8]).replace(" ", "")
                    if phone[0] != '+' and phone[0] != '0':
                        phone = '0' + phone
                    user["phone"] = phone
                except Exception:
                    print(f"[Mauvais format du champ 'phone' pour l'utilisateur '{row[2]}'")
                    has_error = True

            if row[9] is not None:
                user["street"] = row[9]

            if row[10] is not None:
                user["street2"] = row[10]

            if row[11] is not None:
                try:
                    zipcode = str(int(row[11]))
                    if len(zipcode) == 4:
                        zipcode = '0' + zipcode
                    user["zip"] = zipcode
                except Exception:
                    print(f"[Mauvais format du champ 'zipcode' pour l'utilisateur '{row[2]}'")
                    has_error = True
                
            if row[12] is not None:
                user["city"] = row[12]

            if row[13] == "Femme":
                user["sex"]  = "f"
            elif row[13] == "Homme":
                user["sex"]  = "m"

            users.append(user)

    if has_error:
        print("L'import a été interrompu, veuillez régler les erreurs.")
        exit(2)

    res = None
    marshal_none_error = 'cannot marshal None unless allow_none is enabled'
    
    for user in users:
        try:
            res = api.create('res.partner', user)

            if res:
                print(f"Rattaché.e importé.e avec succès : {user['name']} (id : {res})")
        except Exception as e:
            if not (marshal_none_error in str(e)):
                print(f"Erreur lors de l'insertion de {user['name']}, vérifiez ses données dans le tableau ({str(e)})")
            else:
                pass

if __name__ == "__main__":
    main()