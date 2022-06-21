# Create your views here.
from outils.common_imports import *
from outils.for_view_imports import *
from django.views.generic import View
from django.http import HttpResponse
from django.http import JsonResponse

import os
from datetime import date
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from reception.models import CagetteReception
from outils.common import OdooAPI
from outils.common import CouchDB
from members.models import CagetteUser
from products.models import CagetteProduct

# create temp directory if needed
if not os.path.exists("temp"):
    os.mkdir("temp")

def as_text(value):
    """ Utils """
    return str(value) if value is not None else ""

def home(request):
    """Page de selection de la commande suivant un fournisseurs"""
    if 'reception' in settings.COUCHDB['dbs']:
        context = {
            'title': 'Reception',
            'merge_orders_pswd': getattr(settings, 'RECEPTION_MERGE_ORDERS_PSWD', 'makeastop'),
            'couchdb_server': settings.COUCHDB['url'],
            'db': settings.COUCHDB['dbs']['reception'],
            'POUCHDB_VERSION': getattr(settings, 'POUCHDB_VERSION', '')
        }
        template = loader.get_template('reception/index.html')

        return HttpResponse(template.render(context, request))
    else:
        return HttpResponse("Need to configure reception couchdb db in settings_secret.py")


def get_list_orders(request):
    ordersOdoo = CagetteReception.get_orders()
    orders = []
    for order in ordersOdoo:
        # Order with date at 'False' was found.
        try:
            order["date_order"] = time.strftime("%d/%m/%y", time.strptime(order["date_order"], '%Y-%m-%d %H:%M:%S'))
        except:
            pass
        try:
            order["date_planned"] = time.strftime("%d/%m/%y", time.strptime(order["date_planned"], '%Y-%m-%d %H:%M:%S'))
        except:
            pass

        ligne = {
            "id"                : order["id"],
            "name"              : order["name"],
            "date_order"        : order["date_order"],
            "partner"           : order["partner_id"][1],
            "date_planned"      : order["date_planned"],
            "amount_untaxed"    : round(order["amount_untaxed"],2),
            "amount_total"      : round(order["amount_total"],2),
            "reception_status"  : str(order["x_reception_status"])
        }
        orders.append(ligne)

    return JsonResponse({"data": orders}, safe=False)


def produits(request, id):
    """ Gets Order details """
    context = {
        'title': 'Réception des produits',
        "TOOLS_SERVER": settings.TOOLS_SERVER,
        'couchdb_server': settings.COUCHDB['url'],
        'db': settings.COUCHDB['dbs']['reception'],
        'POUCHDB_VERSION': getattr(settings, 'POUCHDB_VERSION', ''),
        "DISPLAY_AUTRES": getattr(settings, 'DISPLAY_COL_AUTRES', True),
        "ADD_ALL_LEFT_IS_GOOD_QTIES": False,
        "ADD_ALL_LEFT_IS_GOOD_PRICES": False,
    }
    fixed_barcode_prefix = '0490'

    if hasattr(settings, 'RECEPTION_PB'):
        context['RECEPTION_PB'] = settings.RECEPTION_PB
    else:
        context['RECEPTION_PB'] = ' [texte à renseigner dans config.py]'

    if hasattr(settings, 'FIXED_BARCODE_PREFIX'):
        fixed_barcode_prefix = settings.FIXED_BARCODE_PREFIX
    if hasattr(settings, 'RECEPTION_ADD_ADMIN_MODE'):
        context['add_admin_mode'] = settings.RECEPTION_ADD_ADMIN_MODE
    if hasattr(settings, 'RECEPTION_ADD_ALL_LEFT_IS_GOOD_QTIES'):
        is_connected_user = CagetteUser.are_credentials_ok(request)
        if is_connected_user is True:
            context['ADD_ALL_LEFT_IS_GOOD_QTIES'] = settings.RECEPTION_ADD_ALL_LEFT_IS_GOOD_QTIES
    if hasattr(settings, 'RECEPTION_ADD_ALL_LEFT_IS_GOOD_PRICES'):
        is_connected_user = CagetteUser.are_credentials_ok(request)
        if is_connected_user is True:
            context['ADD_ALL_LEFT_IS_GOOD_PRICES'] = settings.RECEPTION_ADD_ALL_LEFT_IS_GOOD_PRICES

    context['FIXED_BARCODE_PREFIX'] = fixed_barcode_prefix
    template = loader.get_template('reception/reception_produits.html')

    return HttpResponse(template.render(context, request))


def get_order_lines(request, id_po):
    """Send content of an order"""
    order_lines = CagetteReception.get_order_lines_by_po(int(id_po))

    return JsonResponse({'id_po': id_po, 'po': order_lines})

def get_orders_lines(request):
    """Send content of multiple orders"""
    data = json.loads(request.body.decode())
    orders = []
    for id_po in data['po_ids']:
        order_lines = CagetteReception.get_order_lines_by_po(int(id_po), nullQty = True)
        orders.append({'id_po': id_po, 'po': order_lines})

    return JsonResponse({'orders': orders})

def data_validation(request):
    """ Check if orders can be processed """
    try:
        data = json.loads(request.body.decode())
        unprocessable = []
        for id_po in data:
            order_lines_pid = CagetteReception.get_order_unprocessable_products(int(id_po))
            if len(order_lines_pid) > 0:
                # unprocessable.append({'id_po': id_po, 'products': order_lines_pid})
                unprocessable = unprocessable + order_lines_pid

        return JsonResponse({'unprocessable' : unprocessable})
    except Exception as e:
        coop_logger.error("Orders data validation : %s", str(e))
        return JsonResponse({'error': str(e)}, status=500)

def update_orders(request):
    """Update orders lines: quantity and unit prices"""

    import requests

    # don't print barcode which begin with these codes
    noprint_list = ["0493", "0492", "0499"]

    rep = HttpResponse("Not")
    if request.is_ajax():
        if request.method == 'PUT':
            data = json.loads(request.body.decode())
            if getattr(settings, 'RECEPTION_DATA_BACKUP', True) is True:
                try:
                    file_name = ''
                    for order_id, order in data['orders'].items():
                        file_name += str(order_id) + '_'
                    file_name += data['update_type'] + '_' + str(round(time.time() * 1000)) + '.json'
                    with open('data/receptions_backup/' + file_name, 'w') as data_file:
                        json.dump(data, data_file)
                except Exception as ef:
                    coop_logger.error("Enable to save data : %s (data = %s)",str(ef), str(data))
            answer_data = {}
            print_labels = True
            if hasattr(settings, 'RECEPTION_SHELF_LABEL_PRINT'):
                print_labels = settings.RECEPTION_SHELF_LABEL_PRINT

            order_ids = []
            for order_id, order in data['orders'].items():
                order_ids.append(int(order_id))
                answer_data[order_id] = {}
                errors = []

                m = CagetteReception(order_id)
                try:
                    for order_line in order['po']:
                        if order_line['indicative_package'] is False:
                            m.remove_package_restriction(order_line)

                        update = m.update_line(int(order_line['id']),
                                                data['update_type'],
                                                float(order_line['package_qty']),
                                                float(order_line['product_qty_package']),
                                                float(order_line['price_unit']))
                        if not (update is True):
                            # indicative_package may have been changed since data have been loaded in browser, retry
                            m.remove_package_restriction(order_line)
                            update = m.update_line(int(order_line['id']),
                                                    data['update_type'],
                                                    float(order_line['package_qty']),
                                                    float(order_line['product_qty_package']),
                                                    float(order_line['price_unit']))
                            if not (update is True):
                                errors.append(order_line['id'])

                        # If update succeded, and supplier shortage set, try to register the supplier shortage
                        if update is True and 'supplier_shortage' in order_line:
                            try:
                                answer_data['res_shortage'] = CagetteProduct.register_start_supplier_shortage(
                                                                order_line['product_id'][0],
                                                                order_line['partner_id'][0],
                                                                date.today().strftime("%Y-%m-%d"))
                            except Exception as e:
                                errors.append('error registering shortage on p'+order_line['id']+':'+str(e))

                        # Print etiquette with new price if update if successful and barcode is authorized
                        if (print_labels is True) and (update is True) and (data['update_type'] == 'br_valid') and order_line['new_shelf_price'] and order_line['barcode'][:4] not in noprint_list:
                            try:
                                tools_url = settings.TOOLS_SERVER + '/products/label_print/'
                                tools_url += str(order_line['product_tmpl_id']) + '/'
                                tools_url += str(order_line['new_shelf_price'])
                                requests.get(tools_url)
                            except Exception as e:
                                coop_logger.error("Shelf label printing : %s",str(e))

                except KeyError:
                    coop_logger.info("No line to update.")
                except Exception as e:
                    coop_logger.error("update_orders : %s", str(e))

                answer_data[order_id]['order_data'] = order
                answer_data[order_id]['errors'] = errors

                if len(errors) == 0:
                    m.update_order_status(order_id, data['update_type'])
                    if data['update_type'] == 'br_valid':
                        answer_data[order_id]['finalyze_result'] = m.register_purchase_order_to_finalyze()
                else:
                    coop_logger.error("update_orders errors : %s", str(errors))
            rep = JsonResponse(answer_data, safe=False)
    return rep

# def tmp_update_order_status(request, id_po):
#    """ Method used for tests purposes: Reset an order status """
#    m = CagetteReception(id_po)
#    m.update_order_status(id_po, False)

#    return JsonResponse({'id_po': id_po})

def save_error_report(request):
    """
        Receive json data with the differences between order and reception,
        save it to an excel file,
        attach file to order in odoo
    """
    if request.method == 'POST':
        data = None
        try:
            myJson = request.body
            data = json.loads(myJson.decode())
        except Exception as e1:
            coop_logger.error("Save reception report : Unable to load data %s (%s)", str(e1), str(myJson))

        if data and ('orders' in data):
            orders_name_elts = []
            orders_partner = ""
            group_ids = []

            try:
                for i, order in enumerate(data['orders']) :
                    # list of temp files: 1 report per order & group
                    data['orders'][i]['temp_file_name'] = "temp/" + order['name'] + "_rapport-reception_temp.xlsx"

                    group_ids.append(order['id'])

                    orders_name_elts.append(order['name'])

                    # Concatenate orders partner
                    if order['partner'] != orders_partner:
                        if orders_partner != "":
                            orders_partner = orders_partner + ', '
                        orders_partner = orders_partner + order['partner'] + ' du ' + order['date_order']


                # If group of orders
                if len(data['orders']) > 1:
                    orders_name = '-'.join(orders_name_elts)
                    temp_group_file_name = "temp/" + orders_name + "_rapport-reception_temp.xlsx"

                    # Add group in data['orders']
                    group_order = {
                        'name': orders_name,
                        'partner': orders_partner,
                        'date_order': data['orders'][0]['date_order'],
                        'amount_total': data['group_amount_total'],
                        'updated_products': data['updated_products'],
                        'temp_file_name': temp_group_file_name,
                        'group_ids': group_ids
                    }

                    data['orders'].append(group_order)  # group "order" has to be last in orders list
                else:
                    coop_logger.info("data['orders'] is a single PO (not inside group)")
            except Exception as e2:
                coop_logger.error("Save reception report : Error while create group_order %s", str(e2))

            # no action needed after step 1
            if data['update_type'] == 'qty_valid':
                pass  # removed, keep check to ensure transition process

            # Create report with data from steps 1 & 2
            elif data['update_type'] == 'br_valid':
                c_db = CouchDB(arg_db='reception')

                concat_updated_products = []
                concat_user_comments = ''
                for order in data['orders']:
                    try:
                        data_qties = {}
                        data_comment_s1 = ""

                        if 'group_ids' in order :
                            # For groups, concatenate orders step 1 data 
                            step_1_data = {
                                'updated_products': concat_updated_products,
                                'user_comments': concat_user_comments
                            }
                        else:
                            # Read step 1 data from couch db document
                            order_id = f"order_{order['name'].split('PO')[1]}"
                            order_doc = c_db.getDocById(order_id)

                            try:
                                step_1_data = order_doc['previous_steps_data']['False']
                            except KeyError:
                                continue  # No step 1 data

                        if 'updated_products' in step_1_data:
                            # Concatenate updated products from all orders
                            if 'group_ids' not in order :
                                concat_updated_products = concat_updated_products + step_1_data['updated_products']

                            for product in step_1_data['updated_products']:
                                # Don't store products with unchanged qties in step 1 data
                                if product['old_qty'] != product['product_qty']:
                                    if 'supplier_code' in product:
                                        supplier_code = str(product['supplier_code'])
                                    else:
                                        supplier_code = 'X'

                                    if 'supplier_shortage' in product:
                                        supplier_shortage = '/!\ Rupture fournisseur'
                                    else:
                                        supplier_shortage = ''

                                    product_name = product['product_id'][1]
                                    data_qties[product_name] = {
                                        'nom_contenu' : product_name,
                                        'supplier_code' : supplier_code,
                                        'barcode' : str(product['barcode']),
                                        'old_qty' : product['old_qty'],
                                        'product_qty' : product['product_qty'],
                                        'price_unit' : product['price_unit'],
                                        'supplier_shortage' : supplier_shortage
                                    }

                        if 'user_comments' in step_1_data:
                            # Get user comments from all orders (same for all group orders)
                            if 'group_ids' not in order\
                                and step_1_data['user_comments'] != ""\
                                and concat_user_comments != step_1_data['user_comments']:
                                concat_user_comments = step_1_data['user_comments']

                            data_comment_s1 = step_1_data['user_comments'] if step_1_data['user_comments'] != "" else 'Aucun commentaire.'

                    except Exception as e:
                        data_comment_s1 = "Données de l'étape 1 manquantes (erreur : " + str(e) + ")"

                    # Add data from step 2
                    data_full = []
                    error_total = 0
                    error_total_abs = 0
                    if ('user_comments' in data) and data['user_comments'] != "":
                        data_comment_s2 = data['user_comments']
                    else:
                        data_comment_s2 = "Aucun commentaire."

                    # Concatenate products info from each step
                    try:
                        if "updated_products" in order:
                            for product in order['updated_products']:
                                if 'supplier_code' in product:
                                    supplier_code = str(product['supplier_code'])
                                else:
                                    supplier_code = 'X'

                                if 'supplier_shortage' in product:
                                    supplier_shortage = '/!\ Rupture fournisseur'
                                else:
                                    supplier_shortage = ''

                                item = {
                                    'product_id': product['product_id'][1],
                                    'product_supplier_code': supplier_code,
                                    'product_barcode': product['barcode'],
                                    'old_price_unit': float(product['old_price_unit']),
                                    'price_unit': float(product['price_unit']),
                                    'supplier_shortage': supplier_shortage
                                }

                                # If the product was also modified in step 1
                                if item['product_id'] in data_qties:
                                    item['old_qty'] = float(data_qties[item['product_id']]['old_qty'])
                                    item['product_qty'] = float(data_qties[item['product_id']]['product_qty'])
                                    item['expected_amount'] = item['old_qty']*item['old_price_unit']
                                    item['error_line'] = (item['old_qty'] - item['product_qty'])*item['price_unit']

                                    # If product was set on supplier shortage in step 1 and not in step 2
                                    if item['supplier_shortage'] == '' and data_qties[item['product_id']]['supplier_shortage'] != '':
                                        item['supplier_shortage'] = data_qties[item['product_id']]['supplier_shortage']

                                    data_qties.pop(item['product_id'])
                                else:
                                    item['old_qty'] = float(product['product_qty'])
                                    item['product_qty'] = item['old_qty']
                                    item['expected_amount'] = item['old_qty']*item['old_price_unit']
                                    item['error_line'] = 0
                                    if (item['price_unit'] != item['old_price_unit']):
                                        item['error_line'] = (item['price_unit'] - item['old_price_unit']) * item['product_qty']

                                error_total += item['error_line']
                                error_total_abs += abs(item['error_line'])

                                data_full.append(item)
                        else:
                            coop_logger.info("Save reception error doc : no 'updated_products' in order (%s)", str(order))
                    except Exception as e5:
                        coop_logger.error("Save reception report : Error while updating products %s", str(e5))
                        # no updated products, do nothing

                    # Add remaining products, the ones edited only in step 1
                    for product in data_qties.values():
                        try:
                            item = {
                                'product_id': product['nom_contenu'],
                                'product_supplier_code': product['supplier_code'],
                                'product_barcode': product['barcode'],
                                'old_qty': float(product['old_qty']),
                                'product_qty': float(product['product_qty']),
                                'old_price_unit': float(product['price_unit']),
                                'price_unit': '',
                                'expected_amount':float(product['old_qty'])*float(product['price_unit']),
                                'error_line': (float(product['old_qty'])-float(product['product_qty']))*float(product['price_unit']),
                                'supplier_shortage': product['supplier_shortage']
                            }

                            error_total += item['error_line']
                            error_total_abs += abs(item['error_line'])

                            data_full.append(item)
                        except Exception as e6:
                            coop_logger.error("Save reception report : Error while creating item from product %s (%s)", str(e6), str(e6))
                    try:
                        # Sort by error amount
                        def sortByError(e):
                            return abs(e['error_line'])
                        data_full.sort(reverse=True, key=sortByError)

                        # Create excel file
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Commande " + order['name']
                        # Group
                        if 'group_ids' in order :
                            ws.append( ['Rapport de réception d\'un groupe de commandes'] )
                            ws.append( ['Fournisseur(s) : ', order['partner']] )
                            ws.append( ['Références des commandes du groupe : ', order['name']] )
                        else:
                            ws.append( ['Rapport de réception'] )
                            ws.append( ['Fournisseur : ', order['partner']] )
                            ws.append( ['Réf Commande : ', order['name']] )
                            if len(data['orders']) > 1 :
                                ws.append( ['Commande traitée en groupe. Groupe : ', orders_name] )

                        ws.append( ['Date de la commande : ', order['date_order']] )
                        ws.append( ['Date de la réception : ', time.strftime("%d/%m/%y")] )
                        ws.append( ['Montant total attendu (TTC) : ', str(round(order['amount_total'],2)) + ' €'] )
                        ws.append( [] )

                        ws.append( ['Nom produit',
                                    'Code Four.',
                                    'Numéro de Code Barre',
                                    'Qté commande',
                                    'Qté réception',
                                    'Prix unit. initial',
                                    'Prix unit. MAJ',
                                    'Prix total attendu',
                                    "Montant erreur livraison (basé sur les différences de quantité)"] )

                        if len(data_full) == 0:
                            ws.append( ['- Aucune modification -'] )
                        else:
                            for product in data_full:
                                ws.append( [product['product_id'],
                                            product['product_supplier_code'],
                                            str(product['product_barcode']),
                                            product['old_qty'],
                                            product['product_qty'],
                                            product['old_price_unit'],
                                            product['price_unit'],
                                            round(product['expected_amount'], 2),
                                            round(product['error_line'], 2),
                                            product['supplier_shortage']] )
                        ws.append( [] )
                        ws.append( ['Montant total de l\'erreur :', '', '', '', '', '', '', '', round(error_total, 2)] )
                        ws.append( ['Montant total en valeur absolue :', '', '', '', '', '', '', '', round(error_total_abs, 2)] )

                        ws.append( [] )
                        ws.append( [] )

                        ws.append( ['Problèmes survenus durant le comptage :', data_comment_s1] )
                        # Merge cells for comments
                        merge_begin = ws.max_row
                        merge_end = ws.max_row+3
                        ws.append( [] )
                        ws.append( [] )
                        ws.append( [] )
                        ws.merge_cells(start_row=merge_begin, start_column=1, end_row=merge_end, end_column=1)
                        ws.merge_cells(start_row=merge_begin, start_column=2, end_row=merge_end, end_column=7)
                        # Styling merged cells
                        top_left_cell = ws['A'+str(merge_begin)]
                        top_left_cell.alignment = Alignment(vertical="top")
                        top_left_cell = ws['B'+str(merge_begin)]
                        top_left_cell.alignment = Alignment(vertical="top")

                        ws.append( ['Problèmes survenus durant la mise à jour des prix :', data_comment_s2] )
                        merge_begin = ws.max_row
                        merge_end = ws.max_row+3
                        ws.append( [] )
                        ws.append( [] )
                        ws.append( [] )
                        ws.merge_cells(start_row=merge_begin, start_column=1, end_row=merge_end, end_column=1)
                        ws.merge_cells(start_row=merge_begin, start_column=2, end_row=merge_end, end_column=7)
                        top_left_cell = ws['A'+str(merge_begin)]
                        top_left_cell.alignment = Alignment(vertical="top")
                        top_left_cell = ws['B'+str(merge_begin)]
                        top_left_cell.alignment = Alignment(vertical="top")
                        # "Auto fit" columns width to content
                        for column_cells in ws.columns:
                            length = max(len(as_text(cell.value)) for cell in column_cells)
                            # For other columns than the first, limit size
                            if column_cells[3].column_letter != "A" and length > 20 :
                                length = 20

                            ws.column_dimensions[column_cells[3].column_letter].width = length
                    except Exception as e7:
                        coop_logger.error("PO save report : error while creating final Workbook %s", str(e7))
                    # Save file
                    fileName = "temp/" + order['name'] + "_rapport-reception.xlsx"
                    try:
                        wb.save(filename=fileName)
                        #Attach file to order
                        if 'group_ids' in order:     # group report
                            # Attach group report to each order
                            for group_item_id in order['group_ids']:
                                m = CagetteReception(group_item_id)
                                m.attach_file(fileName, False)
                            os.remove(fileName)
                        else:
                            m = CagetteReception(order['id'])
                            m.attach_file(fileName)
                            coop_logger.info("%s attached to order id %s", fileName, str(order['id']))
                    except Exception as e8:
                        coop_logger.error("PO save report Error while saving file %s (%s)", fileName, str(e8))
            else:
                coop_logger.error("Save reception error report : unknown state %s (%s) ", str(data['update_type']), str(data))
        else:
            coop_logger.error("Cant find 'orders' in data (%s)", str(data))
    else:
        coop_logger.info("Was waiting for a POST method (%s)", str(request.method))
    return JsonResponse("ok", safe=False)

def reception_FAQ(request):
    """Send content of the reception FAQ"""
    context = {}
    template = loader.get_template('reception/reception_FAQ.html')

    return HttpResponse(template.render(context, request))

def reception_qtiesValidated(request):
    """Send content of the validation page after quantities validated"""
    pdt_labels_fn = 'get_pdf_labels()'
    print_text = 'Cliquez sur ce bouton pour télécharger la liste des codes barres à imprimer et à coller sur les produits :'
    print_btn_text = 'Télécharger'

    try:
        pdt_labels_fn = settings.RECEPTION_PDT_LABELS_FN
        print_text = settings.RECEPTION_PDT_LABELS_TEXT
        print_btn_text = settings.RECEPTION_PDT_LABELS_BTN_TEXT

    except:
        pass
    context = {'PRINT_LABELS_TXT': print_text, 'PRINT_BUTTON_TEXT': print_btn_text,
               'GET_PRODUCT_LABELS_FN': pdt_labels_fn}
    template = loader.get_template('reception/reception_modal_qtiesValidated.html')

    return HttpResponse(template.render(context, request))

def reception_pricesValidated(request):
    """Send content of the validation page after prices validated"""
    context = {'RECEPTION_SHELF_LABEL_PRINT': True}
    if hasattr(settings, 'RECEPTION_SHELF_LABEL_PRINT'):
        context['RECEPTION_SHELF_LABEL_PRINT'] = settings.RECEPTION_SHELF_LABEL_PRINT

    template = loader.get_template('reception/reception_modal_pricesValidated.html')

    return HttpResponse(template.render(context, request))

def po_process_picking(request):
    res = CagetteReception.process_enqueued_po_to_finalyze()
    return JsonResponse(res, safe=False)


def send_mail_no_barcode(request):
    """
        Receive json data with liste of product with no barcode
        Send mail to order maker
    """
    from django.core.mail import send_mail

    if request.method == 'POST':
        data = None
        try:
            myJson = request.body
            data = json.loads(myJson.decode())
            data_partner = CagetteReception.get_mail_create_po(int(data['order']['id']))

            msg = settings.NO_BARCODE_MAIL_MSG


            for barcode in data["no_barcode_list"]:

                msg = msg + '       -' + str(barcode[0]) + '---' + str(barcode[1])+ '\n'

            send_mail(settings.NO_BARCODE_MAIL_SUBJECT.format(data['order']['name']),
                  msg.format(data_partner[0]['display_name'], data['order']['name'],data['order']['date_order'], data['order']['partner']),
                  settings.DEFAULT_FROM_EMAIL,
                  [data_partner[0]['email']],
                  fail_silently=False,)


        except Exception as e1:
            coop_logger.error("Send_mail_no_barcode : Unable to load data %s (%s)", str(e1), str(myJson))
            print(str(e1)+'\n'+ str(myJson))


    return JsonResponse("ok", safe=False)
