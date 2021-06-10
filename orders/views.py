from outils.common_imports import *
from outils.for_view_imports import *

from orders.models import Order, Orders, CagetteSuppliers
from products.models import CagetteProduct, CagetteProducts

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook


def as_text(value): return str(value) if value is not None else ""

def index(request):
	return HttpResponse('Orders')

def helper(request):
    context = {
        'title': 'Aide à la commande',
        'couchdb_server': settings.COUCHDB['url'],
        'db': settings.COUCHDB['dbs']['orders']
    }

    template = loader.get_template('orders/helper.html')

    return HttpResponse(template.render(context, request))

def get_suppliers(request):
    """ Get suppliers list """
    res = {}

    try:
        res = CagetteSuppliers.get_suppliers()
    except Exception as e:
        res["error"] = str(e)
        return JsonResponse(res, status=500)

    return JsonResponse({'res': res})

def get_supplier_products(request):
    """ Get supplier products """

    sid = request.GET.get('sid', '')
    res = CagetteProducts.get_products_by_supplier(sid)
    
    if 'error' in res:
        return JsonResponse(res, status=500)
    else:
        return JsonResponse({'res': res})

def associate_supplier_to_product(request):
    """ This product is now supplied by this supplier """
    res = {}
    try:
        data = json.loads(request.body.decode())
        res = CagetteProduct.associate_supplier_to_product(data["product_tmpl_id"], data["supplier_id"])
    except Exception as e:
        res["error"] = str(e)
        return JsonResponse(res, status=500)

    return JsonResponse({'res': res})

def export_one(request, oid):
    msg = ''
    try:
        oid = int(oid)
        order = Order(oid)
        order_data = order.export()
        if ('success' in order_data) and (order_data['success'] is True):
            import datetime
            now = datetime.datetime.now()
            taxes = 0
            company_name = ''
            if hasattr(settings, 'COMPANY_NAME'):
                company_name = settings.COMPANY_NAME

            wb = Workbook()
            ws1 = wb.create_sheet("Commande " + order_data['order']['name'], 0)
            ws1.merge_cells('A1:I1')
            ws1.merge_cells('A2:I2')
            ws1.merge_cells('A3:I3')
            ws1['A1'].value = 'Date : ' + now.strftime("%d/%m/%Y")
            ws1['A2'].value = 'Commande ' + company_name + ' / ' + order_data['order']['partner_id'][1]
            ws1['A3'].value = 'Ref : ' + order_data['order']['name']
            ws1.append([])
            ws1.append(['Produit', 'Nb. colis', 'Colisage',
                        'Qté', 'Référence', 'code-barre','Prix Unitaire', 'Remise', 'Sous-total'])
            for line in order_data['lines']:
                taxes += line['price_tax']
                ws1.append([line['product_id'][1], line['product_qty_package'], line['package_qty'],
                            line['product_qty'], line['supplier_code'], line['barcode'], line['price_unit'], line['discount'], line['price_subtotal']])
            ws1.append([])
            ws1.append(['', '', '', '', '', '', '', 'Montant HT', order_data['order']['amount_untaxed'], 'euros'])
            ws1.append(['', '', '', '', '', '', '', 'Taxes', taxes,  'euros'])
            ws1.append(['', '', '', '', '', '', '', 'Montant TTC', order_data['order']['amount_total'],  'euros'])
            # "Auto fit" columns width to content
            for column_cells in ws1.columns:
                length = max(len(as_text(cell.value)) for cell in column_cells)
                ws1.column_dimensions[column_cells[3].column_letter].width = length

            partner_name = order_data['order']['partner_id'][1]
            partner_name = partner_name.replace("/", "-").replace(" ", "-")
            wb_name = 'commande_' + order_data['order']['name'] + "_" + partner_name + now.strftime("_%Y_%m_%d") + '.xlsx'
            file_path = 'temp/' + wb_name
            wb.save(filename=file_path)
            order.attach_file(file_path)
            msg = 'done'
            # response = HttpResponse(content=save_virtual_workbook(wb),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            # response['Content-Disposition'] = 'attachment; filename=' + wb_name
            # return response
    except Exception as e:
            msg = str(e)

    return JsonResponse({"msg": msg}, safe=False)

def export_regex(request, string):
	return HttpResponse(string + '???')

# Get labels to print for each order (ids received in GET parameter separated by comas)
def get_pdf_labels(request):
    import math
    import io
    from django.http import FileResponse
    msg = ''
    LABELS_PER_SHEET = 65 #  13 rows of 5
    l_data = {'total': 0, 'details': []}

    # Concatenate labels to print for each order received
    order_ids = request.GET['oids'].split(",")
    for id in order_ids:
        order_l_data = Order(id).get_custom_barcode_labels_to_print()
        if order_l_data['total'] > 0:
            l_data['total'] += order_l_data['total']
            l_data['details'] += order_l_data['details']

    if l_data['total'] > 0:
        try:
            from .labels_pdf_generation import pdf_generate
            sheets = []
            labels = 0
            for i in range(0, math.ceil((l_data['total']/LABELS_PER_SHEET))):
                sheets.append([])
            for l in l_data['details']:
                if len(str(l['barcode'])) > 0:
                    # Dispatch all labels to print into sheets
                    nb_before_insert = labels
                    labels += l['product_qty']
                    first_sheet_2_insert = int(nb_before_insert / LABELS_PER_SHEET)
                    first_sheet_places_available = (first_sheet_2_insert + 1) * LABELS_PER_SHEET - nb_before_insert
                    left_to_insert = l['product_qty']
                    ean13 = l['barcode']
                    name = l['product_id'][1]
                    if first_sheet_places_available >= left_to_insert:
                        sheets[first_sheet_2_insert].append({"qty": left_to_insert, "ean13": ean13, "name": name})
                    else:
                        left_to_insert -= first_sheet_places_available
                        sheets[first_sheet_2_insert].append({"qty": first_sheet_places_available, "ean13": ean13, "name": name })
                        sheet_idx = first_sheet_2_insert + 1
                        while left_to_insert > 0:
                            if left_to_insert > 65:
                                qty = 65
                            else:
                                qty = left_to_insert
                            sheets[sheet_idx].append({"qty": qty, "ean13": ean13, "name": name })
                            left_to_insert -= 65
                            sheet_idx += 1
            pdf = pdf_generate(sheets)
            return FileResponse(pdf, as_attachment=True, filename='codebarres.pdf')
        except Exception as e:
            msg = str(e)
    else:
        msg = "Nothing to do !"
    return JsonResponse({"msg": msg}, safe=False)

def print_product_labels(request):
    """Orders ids are given as parameters, to print "own" product labels."""
    res = {}
    oids = []
    try:
        for oid in request.GET['oids'].split(","):
            try:
                oids.append(int(oid))
            except:
                pass  # was not an int
        if len(oids) >= 1:
            ldatas = Orders.get_custom_barcode_labels_to_print(oids)
            for tmpl_id, nb in ldatas.items():
                pres = CagetteProduct.generate_label_for_printing(str(tmpl_id), '/product_labels/' , '0', str(nb))
                if 'error' in pres:
                    if not ('errors' in res):
                        res['errors'] = []
                    res['errors'].append(pres)
    except Exception as e:
        res['error_ext'] = str(e)
    return JsonResponse({'res': res}, safe=False)